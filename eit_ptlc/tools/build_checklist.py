"""Generate plc_l2_checklist.xlsx for acceptance testing."""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ACTIONS = [
    # ── Sampling ──
    ("上样点样", "sampling.init", 10, "上样-初始化", "无"),
    ("上样点样", "sampling.clean", 20, "上样-清洗(注射泵循环)", "无"),
    ("上样点样", "sampling.place_axis", 31, "上样-放板移轴(7Y到放板位)", "无"),
    ("上样点样", "sampling.place_locate", 32, "上样-放板定位(定位夹紧)", "无"),
    ("上样点样", "sampling.prep", 40, "上样-上样准备(Z回零+泵准备)", "无"),
    ("上样点样", "sampling.aspirate", 50, "上样-吸收液体", "plate_spec/plate_no/well, sample_volume_ml"),
    ("上样点样", "sampling.rinse_mix", 55, "上样-加润洗液+抬针吸气隔断+吹打混匀", "plate_spec/plate_no/well, rinse_volume_ml, mix_volume_ml, mix_count, air_gap_ml"),
    ("上样点样", "sampling.spot", 60, "上样-点样", "无"),
    # ── Collect ──
    ("收集", "collect.init", 10, "收集-初始化(复位气缸/阀+泵初始化)", "无"),
    ("收集", "collect.clamp", 21, "收集-夹持气缸夹紧", "无"),
    ("收集", "collect.extend", 22, "收集-伸缩气缸伸出", "无"),
    ("收集", "collect.lift_press", 23, "收集-缩回/升降/下压(缺瓶报错201)", "无"),
    ("收集", "collect.collect", 30, "收集-加液/排液洗脱循环", "无"),
    ("收集", "collect.transport_extend", 41, "收集-复位下压升降+伸缩伸出", "无"),
    ("收集", "collect.retract", 42, "收集-伸缩缩回", "无"),
    ("收集", "collect.release_clamp", 43, "收集-松夹持", "无"),
    # ── Develop ──
    ("展缸", "develop.init", 10, "展缸-初始化", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.clean_line", 20, "展缸-清洗管路(上位机括真空)", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.rinse_fill", 21, "展缸-润洗注液(无真空)", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.rinse_suction", 26, "展缸-润洗抽吸(上位机括真空)", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.fill", 22, "展缸-上液", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.drain", 50, "展缸-排液闭环(Tank_State=99)", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.release_tank", 51, "展缸-取板后释放缸状态", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.plate_retract", 31, "展缸-放板缸回原点", "Expand_Target_Tank:int[1..8]"),
    ("展缸", "develop.plate_extend", 32, "展缸-放板缸到动点", "Expand_Target_Tank:int[1..8]"),
    # ── PhotoScrape ──
    ("拍照刮板", "photoscrape.init", 10, "拍照刮板-初始化", "无"),
    ("拍照刮板", "photoscrape.cam_x335", 31, "拍照-刮板X到放板位335", "无"),
    ("拍照刮板", "photoscrape.cam_locate", 32, "拍照-定位气缸夹紧", "无"),
    ("拍照刮板", "photoscrape.cam_press", 33, "拍照-下压气缸压下", "无"),
    ("拍照刮板", "photoscrape.cam_photopos", 34, "拍照-移相机位+遮光下(就绪待拍)", "无"),
    ("拍照刮板", "photoscrape.cam_photohome", 35, "拍照-遮光上+Y回零", "无"),
    ("拍照刮板", "photoscrape.scrape", 40, "刮取-单pass(g_pass_z由主机预写)", "无"),
    ("拍照刮板", "photoscrape.scrape_finish", 41, "刮取-收尾翻料", "无"),
    ("拍照刮板", "photoscrape.retr_release", 51, "取料-松下压气缸", "无"),
    ("拍照刮板", "photoscrape.retr_stoprot", 52, "取料-停旋转气缸", "无"),
    # ── FeedLift ──
    ("玻璃上下料", "feedlift.feed_raise", 11, "上料-升轴至取料光电(无料报错301)", "无"),
    ("玻璃上下料", "feedlift.feed_lower", 12, "上料-降轴5mm让位", "无"),
    ("玻璃上下料", "feedlift.unload_ready", 21, "下料-到放废料位", "无"),
    ("玻璃上下料", "feedlift.unload_bury", 22, "下料-埋料至光电消失(无料报错302)", "无"),
    # ── Pump (真空泵; 上位机显式括泵, 线圈由 PLC_Pump_泵管理 聚合 大真空泵站位[] 驱动) ──
    ("真空泵", "pump.vacuum_on", 10, "真空泵-开(置泵槽11)", "无"),
    ("真空泵", "pump.vacuum_off", 20, "真空泵-关(清泵槽11)", "无"),
]

STATION_COLORS = {
    "上样点样": "DCE6F1",   # light blue
    "收集":     "E2EFDA",   # light green
    "展缸":     "FDE9D9",   # light orange
    "拍照刮板": "E4DFEC",   # light purple
    "玻璃上下料": "F2DCDB", # light red
    "真空泵":     "FCE4D6", # light orange-2
}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEADERS = ["序号", "工位", "动作名", "动作码", "中文说明", "参数", "测试状态", "问题描述", "备注"]
COL_WIDTHS = [6, 12, 30, 8, 32, 34, 12, 28, 24]

wb = Workbook()
ws = wb.active
ws.title = "PLC L2 验收清单"

# ── Title row ──
ws.merge_cells("A1:I1")
title_cell = ws["A1"]
title_cell.value = "PLC L2 原子动作真机验收检查表"
title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="1F3864")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# Subtitle
ws.merge_cells("A2:I2")
sub = ws["A2"]
sub.value = "测试日期: ___________    测试人: ___________    PLC地址: ___________    控制模式: ___________"
sub.font = Font(name="微软雅黑", size=10, color="555555")
sub.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 24

# ── Header row (row 3) ──
for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[3].height = 28

# ── Data rows ──
prev_station = None
for i, (station, name, code, label, params) in enumerate(ACTIONS):
    row = i + 4
    seq = i + 1
    fill = PatternFill("solid", fgColor=STATION_COLORS.get(station, "FFFFFF"))

    values = [seq, station, name, code, label, params, "待测试", "", ""]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.fill = fill
        if col in (1, 2, 4, 7):  # 序号, 工位, 动作码, 测试状态
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 22

    # Station separator line
    if prev_station and prev_station != station:
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style="thin", color="B0B0B0"),
                right=Side(style="thin", color="B0B0B0"),
                top=Side(style="medium", color="808080"),
                bottom=Side(style="thin", color="B0B0B0"),
            )
    prev_station = station

last_data_row = len(ACTIONS) + 3

# ── Data validation for 测试状态 ──
dv = DataValidation(type="list", formula1='"待测试,通过,失败,跳过,部分通过"', allow_blank=True)
dv.error = "请选择: 待测试/通过/失败/跳过/部分通过"
dv.errorTitle = "无效状态"
ws.add_data_validation(dv)
dv.add(f"G4:G{last_data_row}")

# ── Summary section ──
summ_start = last_data_row + 3
ws.merge_cells(f"A{summ_start}:I{summ_start}")
sc = ws.cell(row=summ_start, column=1, value="统计汇总")
sc.font = Font(name="微软雅黑", bold=True, size=12, color="1F3864")
sc.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[summ_start].height = 26

labels = ["待测试", "通过", "失败", "跳过", "部分通过", "合计"]
for j, label in enumerate(labels):
    r = summ_start + 1 + j
    ws.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=r, column=1).border = THIN_BORDER

    formula = f'=COUNTIF(G4:G{last_data_row},"{label}")' if label != "合计" else f"=SUM(B{summ_start+1}:B{summ_start+5})"
    cell = ws.cell(row=r, column=2, value=formula)
    cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.fill = PatternFill("solid", fgColor="F2F2F2")
    cell.border = THIN_BORDER

# Progress bar
progress_r = summ_start + 1
ws.merge_cells(f"C{progress_r}:I{progress_r}")
pct_cell = ws.cell(row=progress_r, column=3, value=f'=IF(B{summ_start+6}>0, B{summ_start+2}/B{summ_start+6}, 0)')
pct_cell.number_format = '0.0%'
pct_cell.alignment = Alignment(horizontal="center", vertical="center")
pct_cell.font = Font(name="微软雅黑", size=18, bold=True, color="1F3864")
# Merge progress rows for visual impact
for j in range(6):
    ws.merge_cells(f"C{summ_start+1+j}:I{summ_start+1+j}")
for j in range(6):
    c = ws.cell(row=summ_start+1+j, column=3)
    c.border = THIN_BORDER
    c.font = BOLD_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    if j == 0:
        c.value = f'="测试通过率: "&TEXT(IF(B{summ_start+6}>0, B{summ_start+2}/B{summ_start+6}, 0),"0.0%")&"  ("&B{summ_start+2}&"/"&B{summ_start+6}&")"'
    elif j == 1:
        c.value = f'="测试完成率: "&TEXT(IF(B{summ_start+6}>0, (B{summ_start+6}-B{summ_start+1})/B{summ_start+6}, 0),"0.0%")&"  ("&(f"B{summ_start+6}-B{summ_start+1}")&"/"&B{summ_start+6}&")"'
        # Fix: use formula for tested count
        c.value = f'="测试完成率: "&TEXT(IF(B{summ_start+6}>0, (B{summ_start+2}+B{summ_start+3}+B{summ_start+5})/B{summ_start+6}, 0),"0.0%")'
    elif j == 2:
        c.value = f'="失败率: "&TEXT(IF(B{summ_start+6}>0, B{summ_start+3}/B{summ_start+6},0),"0.0%")'
    elif j == 3:
        c.value = ""
    elif j == 4:
        c.value = ""
    elif j == 5:
        c.value = ""

# ── Legend ──
legend_r = summ_start + 9
ws.merge_cells(f"A{legend_r}:I{legend_r}")
ws.cell(row=legend_r, column=1, value="工位颜色图例:").font = Font(name="微软雅黑", bold=True, size=10)

for k, (st, color) in enumerate(STATION_COLORS.items()):
    c = ws.cell(row=legend_r + 1, column=1 + k * 2)
    c.value = st
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(name="微软雅黑", size=9)
    c.border = THIN_BORDER
    c.alignment = CENTER

# ── Freeze & filter ──
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:I{last_data_row}"

# ── Print settings ──
ws.sheet_properties.pageSetUpPr = None
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

OUT = __file__.replace("build_checklist.py", "plc_l2_checklist.xlsx")
wb.save(OUT)
print(f"Saved → {OUT}")
